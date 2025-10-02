from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from typing import Any

try:  # pragma: no cover - optional dependency guard
    import openpyxl  # type: ignore
except Exception:  # noqa: BLE001
    openpyxl = None  # type: ignore

from .base_types import RawRow, UnsupportedFormatError

__all__ = ["XLSXImportResult", "parse_xlsx"]


@dataclass(slots=True)
class XLSXImportResult:
    rows: list[RawRow]
    headers: Sequence[str]


def parse_xlsx(data: bytes) -> XLSXImportResult:
    """Parse an XLSX file (first sheet) into raw rows.

    Rules mirror CSV/DOCX importers:
    - First row considered headers.
    - Blank rows skipped (all empty or None after str+strip).
    - Short rows padded, long rows truncated to header length.
    """
    if openpyxl is None:  # pragma: no cover - environment branch
        raise UnsupportedFormatError("openpyxl is not installed")

    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)  # type: ignore[attr-defined]
    if not wb.sheetnames:
        return XLSXImportResult(rows=[], headers=[])
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return XLSXImportResult(rows=[], headers=[])

    headers = [(_cell_to_str(c)) for c in raw_headers]

    rows: list[RawRow] = []
    for r in rows_iter:
        values = [(_cell_to_str(c)) for c in r[: len(headers)]]
        if all(v == "" for v in values):
            continue
        # pad if shorter
        if len(values) < len(headers):
            values.extend(["" for _ in range(len(headers) - len(values))])
        mapped: RawRow = {}
        for h, v in zip(headers, values, strict=False):
            mapped[h] = v
        rows.append(mapped)

    return XLSXImportResult(rows=rows, headers=headers)


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().strip("\ufeff")
