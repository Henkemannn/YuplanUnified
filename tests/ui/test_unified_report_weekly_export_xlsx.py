from sqlalchemy import text

HEADERS = {"X-User-Role": "admin", "X-Tenant-Id": "1"}


def _seed_site_and_department():
    from core.db import get_session
    conn = get_session()
    try:
        site = conn.execute(text("SELECT id FROM sites WHERE id='00000000-0000-0000-0000-000000000000'"))\
            .fetchone()
        if not site:
            conn.execute(text("INSERT INTO sites (id, name) VALUES ('00000000-0000-0000-0000-000000000000', 'Test Site')"))
        dep = conn.execute(text("SELECT id FROM departments WHERE id='00000000-0000-0000-0000-000000000001'"))\
            .fetchone()
        if not dep:
            conn.execute(text("INSERT INTO departments (id, site_id, name, resident_count_mode, resident_count_fixed) VALUES ('00000000-0000-0000-0000-000000000001', '00000000-0000-0000-0000-000000000000', 'Avd Alpha', 'fixed', 5)"))
        conn.commit()
    finally:
        conn.close()


def test_weekly_report_xlsx_happy_path(app_session):
    client = app_session.test_client()
    _seed_site_and_department()

    site_id = "00000000-0000-0000-0000-000000000000"
    year = 2025
    week = 48

    resp_xlsx = client.get(
        f"/ui/reports/weekly.xlsx?site_id={site_id}&year={year}&week={week}",
        headers=HEADERS,
    )
    assert resp_xlsx.status_code == 200
    assert resp_xlsx.headers.get("Content-Type", "").startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    cd = resp_xlsx.headers.get("Content-Disposition", "")
    assert f"veckorapport_v{week}_{year}.xlsx" in cd

    # Read workbook from response
    from io import BytesIO
    buf = BytesIO(resp_xlsx.data)
    try:
        from openpyxl import load_workbook
    except Exception:
        assert False, "openpyxl not available for XLSX test"
    wb = load_workbook(buf)
    ws = wb.active

    # Header row check (must match labels used in exporter)
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == [
        "Site",
        "Avdelning",
        "År",
        "Vecka",
        "Måltid",
        "Boende totalt",
        "Gjorda specialkoster",
        "Normalkost",
    ]

    # There should be at least two data rows (lunch/dinner for the department)
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) >= 2
    # Ensure debiterbar specialkost and normal count columns exist and are numeric
    for row in data_rows:
        residents_total = row[5] or 0
        debiterbar = row[6] or 0
        normal = row[7] or 0
        assert isinstance(debiterbar, (int, float))
        assert isinstance(normal, (int, float))
        # Normal should equal residents_total - debiterbar (non-negative)
        assert normal == max(0, (residents_total or 0) - (debiterbar or 0))
